# -*- coding: utf-8 -*-
"""WULIX - Generateur template Excel suivi revenus"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import datetime

OUTPUT = r"C:\Users\USER\.claude\projects\projet jarvis\agents\content\WULIX_Suivi_Revenus.xlsx"

wb = openpyxl.Workbook()

# Couleurs
PURPLE = "7C3AED"
CYAN   = "00E5FF"
DARK   = "0D0020"
GRAY   = "CCCCCC"
WHITE  = "FFFFFF"
GREEN  = "22C55E"
YELLOW = "EAB308"

def style_header(cell, bg=PURPLE, fg=WHITE, bold=True, size=11):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=fg, bold=bold, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def border_cell(cell):
    thin = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


# ============================================================
# ONGLET 1 — Gumroad
# ============================================================
ws1 = wb.active
ws1.title = "Gumroad"

ws1["A1"] = "WULIX — Suivi Ventes Gumroad"
ws1["A1"].font = Font(color=CYAN, bold=True, size=14)
ws1["A1"].fill = PatternFill("solid", fgColor=DARK)
ws1.merge_cells("A1:G1")
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

headers = ["Date", "Produit", "Prix (EUR)", "Commission Gumroad (10%)", "Net Recu (EUR)", "Email acheteur", "Notes"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=2, column=col, value=h)
    style_header(c)
    ws1.column_dimensions[c.column_letter].width = 20

# Lignes exemple
exemples = [
    [datetime.date(2026, 4, 18), "Pack Scripts Python", 29, "=C3*0.1", "=C3-D3", "client@example.com", "Premiere vente"],
    [datetime.date(2026, 4, 19), "Guide PDF 9EUR",       9,  "=C4*0.1", "=C4-D4", "", ""],
    [datetime.date(2026, 4, 20), "Pipeline LinkedIn",   19,  "=C5*0.1", "=C5-D5", "", ""],
]
for row_idx, row_data in enumerate(exemples, 3):
    for col_idx, val in enumerate(row_data, 1):
        c = ws1.cell(row=row_idx, column=col_idx, value=val)
        border_cell(c)
        if col_idx in (3, 4, 5):
            c.number_format = '#,##0.00 "EUR"'
        c.alignment = Alignment(vertical="center")

# Ligne total
total_row = 7
ws1.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color=WHITE)
ws1.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=PURPLE)
ws1.cell(row=total_row, column=3, value="=SUM(C3:C6)").number_format = '#,##0.00 "EUR"'
ws1.cell(row=total_row, column=5, value="=SUM(E3:E6)").number_format = '#,##0.00 "EUR"'
for col in (3, 5):
    c = ws1.cell(row=total_row, column=col)
    c.fill = PatternFill("solid", fgColor=GREEN)
    c.font = Font(bold=True, color=WHITE)

ws1.row_dimensions[2].height = 22


# ============================================================
# ONGLET 2 — Fiverr
# ============================================================
ws2 = wb.create_sheet("Fiverr")
ws2["A1"] = "WULIX — Suivi Missions Fiverr"
ws2["A1"].font = Font(color=CYAN, bold=True, size=14)
ws2["A1"].fill = PatternFill("solid", fgColor=DARK)
ws2.merge_cells("A1:G1")
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

headers2 = ["Date", "Client", "Description mission", "Prix brut (EUR)", "Commission Fiverr (20%)", "Net Recu (EUR)", "Statut"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=2, column=col, value=h)
    style_header(c)
    ws2.column_dimensions[c.column_letter].width = 22

ex2 = [
    [datetime.date(2026, 4, 15), "client_us", "Script automatisation Python", 150, "=D3*0.2", "=D3-E3", "Livre"],
    [datetime.date(2026, 4, 17), "client_fr", "Agent IA scraping", 75,  "=D4*0.2", "=D4-E4", "En cours"],
]
for ri, row_data in enumerate(ex2, 3):
    for ci, val in enumerate(row_data, 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        border_cell(c)
        if ci in (4, 5, 6):
            c.number_format = '#,##0.00 "EUR"'


# ============================================================
# ONGLET 3 — Malt
# ============================================================
ws3 = wb.create_sheet("Malt")
ws3["A1"] = "WULIX — Suivi Missions Malt"
ws3["A1"].font = Font(color=CYAN, bold=True, size=14)
ws3["A1"].fill = PatternFill("solid", fgColor=DARK)
ws3.merge_cells("A1:G1")
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

headers3 = ["Mois", "Client", "Nb jours", "TJM (EUR)", "Brut (EUR)", "Commission Malt (10%)", "Net (EUR)"]
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=2, column=col, value=h)
    style_header(c, bg="1A0030")
    ws3.column_dimensions[c.column_letter].width = 20

ex3 = [["Avril 2026", "Client A", 2, 350, "=C3*D3", "=E3*0.1", "=E3-F3"]]
for ri, row_data in enumerate(ex3, 3):
    for ci, val in enumerate(row_data, 1):
        c = ws3.cell(row=ri, column=ci, value=val)
        border_cell(c)
        if ci in (4, 5, 6, 7):
            c.number_format = '#,##0.00 "EUR"'


# ============================================================
# ONGLET 4 — Recap mensuel
# ============================================================
ws4 = wb.create_sheet("Recap Mensuel")
ws4["A1"] = "WULIX — Recap Mensuel Consolide"
ws4["A1"].font = Font(color=CYAN, bold=True, size=14)
ws4["A1"].fill = PatternFill("solid", fgColor=DARK)
ws4.merge_cells("A1:F1")
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

headers4 = ["Mois", "Gumroad (EUR)", "Fiverr (EUR)", "Malt (EUR)", "Total (EUR)", "Objectif (EUR)"]
for col, h in enumerate(headers4, 1):
    c = ws4.cell(row=2, column=col, value=h)
    style_header(c, bg="1A0030")
    ws4.column_dimensions[c.column_letter].width = 18

mois = ["Avril 2026", "Mai 2026", "Juin 2026", "Juillet 2026", "Aout 2026", "Septembre 2026"]
objectifs = [500, 800, 1200, 1500, 2000, 2500]
for i, (m, obj) in enumerate(zip(mois, objectifs), 3):
    ws4.cell(row=i, column=1, value=m)
    ws4.cell(row=i, column=2, value=0).number_format = '#,##0.00 "EUR"'
    ws4.cell(row=i, column=3, value=0).number_format = '#,##0.00 "EUR"'
    ws4.cell(row=i, column=4, value=0).number_format = '#,##0.00 "EUR"'
    total_cell = ws4.cell(row=i, column=5, value=f"=B{i}+C{i}+D{i}")
    total_cell.number_format = '#,##0.00 "EUR"'
    total_cell.font = Font(bold=True, color=GREEN)
    ws4.cell(row=i, column=6, value=obj).number_format = '#,##0.00 "EUR"'
    for col in range(1, 7):
        border_cell(ws4.cell(row=i, column=col))

# Graphique
chart = BarChart()
chart.title = "Revenus vs Objectifs"
chart.style = 10
chart.y_axis.title = "EUR"
chart.x_axis.title = "Mois"
chart.width = 20
chart.height = 12

data = Reference(ws4, min_col=5, max_col=6, min_row=2, max_row=8)
cats = Reference(ws4, min_col=1, min_row=3, max_row=8)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws4.add_chart(chart, "A12")


# ============================================================
# ONGLET 5 — Objectifs
# ============================================================
ws5 = wb.create_sheet("Objectifs")
ws5["A1"] = "WULIX — Objectifs & KPIs"
ws5["A1"].font = Font(color=CYAN, bold=True, size=14)
ws5["A1"].fill = PatternFill("solid", fgColor=DARK)
ws5.merge_cells("A1:C1")
ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 30

kpis = [
    ("KPI", "Objectif", "Statut"),
    ("Revenu mensuel Avril 2026", "500 EUR", "En cours"),
    ("Revenu mensuel Mai 2026", "800 EUR", "A venir"),
    ("Ventes Gumroad Produit 1", "5 ventes/mois", "A venir"),
    ("Ventes Gumroad Produit 2", "3 ventes/mois", "A venir"),
    ("Ventes Gumroad Produit 3", "8 ventes/mois", "A venir"),
    ("Missions Fiverr", "2 missions/mois", "En cours"),
    ("Missions Malt", "1 mission/mois (350EUR/j)", "En cours"),
    ("Profil Malt completude", "100% (photo manquante)", "En cours"),
    ("Google Search Console", "Soumettre sitemap", "A faire"),
    ("Abonnes LinkedIn", "+100/mois", "A venir"),
]
for ri, row_data in enumerate(kpis, 2):
    for ci, val in enumerate(row_data, 1):
        c = ws5.cell(row=ri, column=ci, value=val)
        ws5.column_dimensions[c.column_letter].width = 35
        border_cell(c)
        if ri == 2:
            style_header(c)
        elif val in ("En cours",):
            c.fill = PatternFill("solid", fgColor="EAB308")
            c.font = Font(bold=True, color=WHITE)
        elif val in ("A faire",):
            c.fill = PatternFill("solid", fgColor="EF4444")
            c.font = Font(bold=True, color=WHITE)
        elif val in ("A venir",):
            c.fill = PatternFill("solid", fgColor="6B7280")
            c.font = Font(bold=True, color=WHITE)
        elif val in ("Fait",):
            c.fill = PatternFill("solid", fgColor=GREEN)
            c.font = Font(bold=True, color=WHITE)

wb.save(OUTPUT)
print("[OK] Excel genere :", OUTPUT)
print("     Onglets : Gumroad | Fiverr | Malt | Recap Mensuel | Objectifs")
